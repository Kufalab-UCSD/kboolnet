#!/usr/bin/python3

import logging
import os
import sys
import re

import click
import click_log
import colorama

from rxncon.input.excel_book.excel_book import ExcelBook
from rxncon.visualization.regulatory_graph import SpeciesReactionGraph
from rxncon.visualization.graphML import XGMML
from rxncon.visualization.graphML import map_layout2xgmml

import networkx as nx
from networkx import DiGraph

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

colorama.init()


def extract_modules(excel_filename: str, output=None, modules=[], min_quality=0):
    """
    Extracts modules from an excel input, adds any reactions necessary to satisfy said contingency,
    and then writes ouput to a new file

    Args:
        excel_filename: Name of the excel input file.
        output: Name of the new output.
        modules: Comma-separated list of modules to be extracted
        min_quality: Minimum quality for a line to be kept
    """
    logger.debug("modules: {}".format(modules))
    logger.debug("min quality: {}".format(min_quality))

    if not output:
        output = os.path.dirname(excel_filename) + 'modules.xlsx'

    if not output.endswith('.xlsx'):
        output +=  '.xlsx'

    print('Reading in Excel file [{}] ...'.format(excel_filename))
    excel_book = ExcelBook(excel_filename)

    rxncon_system = excel_book.rxncon_system
    print('Constructed rxncon system: [{} reactions], [{} contingencies]'
          .format(len(rxncon_system.reactions), len(rxncon_system.contingencies)))

    print('Generating regulatory graph output...')
    reg_system = SpeciesReactionGraph(rxncon_system)
    graph = reg_system.to_graph()

    ### STEP 1: Get all the reactions and contingencies and filter down to only those with module tag + minimum quality
    # Get the sheets
    con_sheet = excel_book._xlrd_book.sheet_by_name('ContingencyList')
    rxn_sheet = excel_book._xlrd_book.sheet_by_name('ReactionList')

    # Get module and quality column numbers for each sheet
    rxn_header_row = list(rxn_sheet.get_rows())[1]
    for num, header in enumerate(rxn_header_row):
        if header.value == '!Module':
            column_rxn_module = num
        elif header.value == '!Quality':
            column_rxn_quality = num
    
    con_header_row = list(con_sheet.get_rows())[1]
    for num, header in enumerate(con_header_row):
        if header.value == '!Module':
            column_con_module = num
        elif header.value == '!Quality':
            column_con_quality = num

    # Filter rxn rows by module tag and min quality
    rxn_rows = [row for row in rxn_sheet.get_rows()][2:]
    rxn_filtered = []

    for num, row in enumerate(rxn_rows):
        row_modules = list(map(str.strip, row[column_rxn_module].value.split(','))) # Split modules into list and trim whitespace
        
        if not row[column_rxn_quality].value == '': # If quality missing, keep the row
            try: # Try and parse quality to an int
                row_quality = int(row[column_rxn_quality].value)
            except ValueError:
                raise SyntaxError('Reaction {}\'s quality must be an int.'.format(row[excel_book._column_reaction_full_name].value))
            
            if row_quality < min_quality: # If less than minimum quality skip row
                logger.debug('Skipping reaction row {} (reaction quality: {}, min quality: {})'.format(num, row_quality, min_quality))
                continue
        elif not modules == [] and not any(item in row_modules for item in modules): # If row module isnt in modules list, skip row
            logger.debug('Skipping reaction row {} (module not present)'.format(num))
            continue
        else:
            rxn_filtered.append({'row_num': num, 'name': row[excel_book._column_reaction_full_name].value})
        
    # Filter rxn rows by module tag and min quality
    con_rows = [row for row in con_sheet.get_rows()][2:]
    con_filtered = []

    for num, row in enumerate(con_rows):
        row_modules = list(map(str.strip, row[column_con_module].value.split(','))) # Split modules into list and trim whitespace
        
        if not row[column_con_quality].value == '': # If quality missing, keep the row
            try: # Try and parse quality to an int
                row_quality = int(row[column_con_quality].value)
            except ValueError:
                raise SyntaxError('Contingency {}\'s quality must be an int.'.format(num))
            
            if row_quality < min_quality: # If less than minimum quality skip row
                logger.debug('Skipping contingency row {} (reaction quality: {}, min quality: {})'.format(num, row_quality, min_quality))
                continue
        elif not modules == [] and not any(item in row_modules for item in modules): # If row module isnt in modules list, skip row
            logger.debug('Skipping contingency row {} (module not present)'.format(num))
            continue
        else:
            con_filtered.append({'row_num': num, 'target': row[excel_book._column_contingency_target].value.strip(),
                                 'modifier': row[excel_book._column_contingency_modifier].value.strip()})

    ### STEP 2: Map contingencies to nodes
    required_components = []
    for con in con_filtered:
        row = con_rows[con['row_num']]
        con['node'] = ''

        if con['modifier'] in graph.node: # If modifier directly found in graph
            con['node'] = con['modifier']
        if re.search(r'.*\[.*\]--.*\[.*\](?!0$)', con['modifier']): # If modifier is a complex (with domains)
            components = con['modifier'].split('--')
            con['node'] = components[0] + '--' + components[1]
            if not con['node'] in graph.nodes: # Try and reverse the complex if that didn't work
                con['node'] = components[1] + '--' + components[0]
        elif re.search(r'[A-Za-z0-9]+--[A-Za-z0-9]+', con['modifier']): # If modifier is a complex (no domains)
            components = con['modifier'].split('--')
            r = re.compile(components[0] + r'_\[.*?\]--' + components[1] + r'_\[.*?\]') # Try and modifier with domains
            matches = list(filter(r.match, list(graph.nodes.keys())))

            if len(matches) == 1:
                con['node'] = matches[0]
            elif len(matches) == 0: # If that fails, try the other order
                r = re.compile(components[1] + r'_\[.*?\]--' + components[0] + r'_\[.*?\]') 
                matches = list(filter(r.match, list(graph.nodes.keys())))

                if len(matches) == 1:
                    con['node'] = matches[0]
        elif re.search(r'<.*>', con['modifier']): # If modifier is a boolean node
            con['node'] = con['modifier'][1:len(con['modifier'])-1] # Trim out the brackets
        elif re.search(r'[A-Za-z0-9]+-{.*}', con['modifier']): # If modifier is a modification (no domains)
            components = con['modifier'].split('-')
            r = re.compile(components[0] + r'_\[.*\]-' + components[1], re.IGNORECASE)
            matches = list(filter(r.match, list(graph.nodes.keys())))

            if len(matches) == 1:
                con['node'] = matches[0]
        else: # Try case insensitive search
            r = re.compile(re.escape(con['modifier']), re.IGNORECASE)
            matches = list(filter(r.match, list(graph.nodes.keys())))

            if len(matches) == 1:
                con['node'] = matches[0]

        if not con['node'] in graph.node:
            raise ValueError('Unable to match contingency {} to a node in the srgraph.'.format(con['modifier']))
        else:
            if not re.search(r'^(<.*>|\[.*\])$', con['modifier']): # If not a boolean node or a global state
                components = re.search(r'(?:([A-Za-z0-9]*)_\[.*?\](?:.*?([A-Za-z0-9]*)_\[.*?\])?|^[A-Za-z0-9]*$)', con['node']).groups() # Get all components in the state
                components = [component for component in components if component != None] # Discard None components
                required_components.extend(components) # Append components for state to list of required components
                logger.debug('Matched contingency {} to node {}, contains components {}'.format(con['modifier'], con['node'], components))
            else: # If a boolean node or global state
                logger.debug('Matched contingency {} to node {}'.format(con['modifier'], con['node']))

    ### STEP 3: For each filtered contingency (ie a state), add reaction(s) that produce(s) state to list (also add target reactions)
    required_components = list(set(required_components)) # Make sure there are no duplicate components
    required_reactions = []
    for con in con_filtered:
        if not re.search(r'^(<.*>|\[.*\])$', con['target']): # If target not a global state or boolean node, add it
            required_reactions.append(con['target'])

        if graph.nodes(data='type')[con['node']] == 'state':
            con_reactions = [r for r in graph.predecessors(con['node'])]
            con_required_reactions = []

            for reaction in con_reactions: # For all of the reactions that produce that state
                reaction_components = re.search(r'([A-Za-z0-9]*)(?:_\[.*\])?_.*?_([A-Za-z0-9]*)(?:_\[.*\])?', reaction).group(1,2) # Get both the components involved in the reaction
                if reaction_components[0] in required_components and reaction_components[1] in required_components:
                    con_required_reactions.append(reaction) # If both components of reaction are required components, add to con_required_reactions

            if len(con_required_reactions) > 0: # If reaction producing state that only involves require components found, add it to required reactions
                required_reactions.extend(con_required_reactions)
            else: # Otherwise just add all the reactions that produce that state
                required_reactions.extend(con_reactions)
    required_reactions = list(dict.fromkeys(required_reactions))

    ### STEP 4: Check if required reactions already in rxn_filtered and add them if missing
    all_rxns = rxn_sheet.col_values(excel_book._column_reaction_full_name)[2:] # Get list of all reaction names
    rxn_names = [d['name'] for d in rxn_filtered]
    for node in required_reactions:
        bidirectional = re.sub(r'[+-](?=_.*?)', '', node)
        no_domain = re.sub(r'_\[.*?\]', '', node)
        no_domain_bidirectional = re.sub(r'_\[.*?\]', '', bidirectional)

        # Check if node exists in rxn_filtered and skip it if it does
        if node in rxn_names or no_domain in rxn_names or bidirectional in rxn_names or no_domain_bidirectional in rxn_names:
            continue

        # If node not in rxn_filtered, try and find the matching reaction and add it to rxn_filtered
        for num, rxn in enumerate(all_rxns):
            if node == rxn or bidirectional == rxn or no_domain == rxn or no_domain_bidirectional == rxn:
                logger.warning('Added required reaction {} to filtered reactions.'.format(rxn, node))
                rxn_filtered.append({'name': rxn, 'row_num': num})
                break

            if num == len(all_rxns)-1:
                raise ValueError('Unable to find reaction that matches node {}'.format(node))

    ### STEP 5: Write updated lists to excel file. This has to be done with openpyxl as xlrd/xlwt don't really support xlsx
    # NOTE: openpyxl is 1-indexed, not 0-indexed like xlrd
    wb = load_workbook(filename = excel_filename, data_only = True)
    rxn_open_sheet = wb['ReactionList']
    con_open_sheet = wb['ContingencyList']

    # Delete the old unfiltered rows. This is more efficient than the delete_rows method
    for row in rxn_open_sheet.iter_rows(min_row=3, max_row=(3+len(rxn_rows))):
        for col in range(0,12):
            row[col].value = None

    for row in con_open_sheet.iter_rows(min_row=3, max_row=(3+len(con_rows))):
        for col in range(0,8):
            row[col].value = None 
    

    # Add the filtered rows to workbook
    rxn_filtered_row_nums = [d['row_num'] for d in rxn_filtered]
    for row_num, filtered_row_num in enumerate(rxn_filtered_row_nums, 3):
        for col_num, val in enumerate(rxn_sheet.row_values(filtered_row_num+2), 1):
            rxn_open_sheet.cell(row_num, col_num).value = val

    con_filtered_row_nums = [d['row_num'] for d in con_filtered]
    for row_num, filtered_row_num in enumerate(con_filtered_row_nums, 3):
        for col_num, val in enumerate(con_sheet.row_values(filtered_row_num+2), 1):
            con_open_sheet.cell(row_num, col_num).value = val

    # Write to file
    wb.save(output)
    print('Wrote filtered rxncon file to {}'.format(output))



@click.command()
@click.option('--file', required=True,
              help='Name of input file.')
@click.option('--output', default=None,
              help='Name of output file. Default: modules.xlsx')
@click.option('--modules', default='',
              help='Comma-separated list of modules to be extracted from input file.')
@click.option('--quality', default=0, type=click.INT,
              help='Minimum quality of rules to be kept. Default: 0')
@click_log.simple_verbosity_option(default='WARNING')
@click_log.init()
def run(file, output, modules, quality):
    modules = list(map(str.strip, modules.split(','))) # Split modules into list and trim whitespace
    modules = list(filter(None, modules)) # Filter out blank modules
    extract_modules(file, output, modules, quality)


def setup_logging_colors():
    click_log.ColorFormatter.colors = {
        'error': dict(fg='red'),
        'exception': dict(fg='red'),
        'critical': dict(fg='red'),
        'debug': dict(fg='yellow'),
        'warning': dict(fg='yellow'),
        'info': dict(fg='yellow')
    }

    def format(self, record):
        if not record.exc_info:
            level = record.levelname.lower()
            if level in self.colors:
                padding_size = 7  # Assume just INFO / DEBUG entries.

                prefix = click.style('{}: '.format(level).ljust(padding_size),
                                     **self.colors[level])

                prefix += click.style('{} '.format(record.name), fg='blue')

                msg = record.msg
                if isinstance(msg, bytes):
                    msg = msg.decode(sys.getfilesystemencoding(),
                                     'replace')
                elif not isinstance(msg, str):
                    msg = str(msg)
                record.msg = '\n'.join(prefix + x for x in msg.splitlines())

        return logging.Formatter.format(self, record)

    click_log.ColorFormatter.format = format


if __name__ == '__main__':
    # try:
    setup_logging_colors()
    run()
    # except Exception as e:
    #     print('ERROR: {}\n{}\nPlease re-run this command with the \'-v DEBUG\' option.'.format(type(e), e))

