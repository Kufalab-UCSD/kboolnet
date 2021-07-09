#!/usr/bin/python3

import logging
import os
import sys
import re

import click
import click_log
import colorama

from rxncon.input.excel_book.excel_book import ExcelBook

from rxncon.core.state import state_from_str, GLOBAL_STATE_REGEX
from rxncon.core.effector import BOOLEAN_CONTINGENCY_REGEX
from rxncon.core.reaction import BIDIRECTIONAL_REACTIONS

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

colorama.init()


def extract_modules(excel_filename: str, output=None, modules=[], min_quality=0):
    """
    Extracts modules from an excel input, adds any reactions necessary to satisfy said contingency,
    and then writes ouput to a new file

    Args:
        excel_filename: Name of the excel input file.
        output: Name of the new output.
        modules: Comma-separated list of modules to be extracted
        min_quality: Minimum quality for a line to be kept
    """
    logger.debug("modules: {}".format(modules))
    logger.debug("min quality: {}".format(min_quality))

    if not output:
        output = os.path.dirname(excel_filename) + 'modules.xlsx'

    if not output.endswith('.xlsx'):
        output +=  '.xlsx'

    print('Reading in Excel file [{}] ...'.format(excel_filename))
    excel_book = ExcelBook(excel_filename)

    rxncon_system = excel_book.rxncon_system
    print('Constructed rxncon system: [{} reactions], [{} contingencies]'
          .format(len(rxncon_system.reactions), len(rxncon_system.contingencies)))

    ### STEP 1: Get all the reactions and contingencies and filter down to only those with module tag + minimum quality
    # Get the sheets
    con_sheet = excel_book._xlrd_book.sheet_by_name('ContingencyList')
    rxn_sheet = excel_book._xlrd_book.sheet_by_name('ReactionList')

    # Get module and quality column numbers for reactions
    column_rxn_module = None
    column_rxn_quality = None
    rxn_header_row = list(rxn_sheet.get_rows())[1]
    for num, header in enumerate(rxn_header_row):
        if header.value == '!Module':
            column_rxn_module = num
        elif header.value == '!Quality':
            column_rxn_quality = num

    if column_rxn_module is None:
        raise ValueError('You must define a !Module column in the ReactionList sheet.')
    elif column_rxn_quality is None:
        raise ValueError('You must define a !Quality column in the ReactionList sheet.')
    
    # Get module and quality column numbers for contingencies
    column_con_module = None
    column_con_quality = None
    con_header_row = list(con_sheet.get_rows())[1]
    for num, header in enumerate(con_header_row):
        if header.value == '!Module':
            column_con_module = num
        elif header.value == '!Quality':
            column_con_quality = num

    if column_con_module is None:
        raise ValueError('You must define a !Module column in the ContingencyList sheet.')
    elif column_con_quality is None:
        raise ValueError('You must define a !Quality column in the ContingencyList sheet.')

    # Filter rxn rows by module tag and min quality
    rxn_rows = [row for row in rxn_sheet.get_rows()][2:]
    rxn_filtered = []

    for num, row in enumerate(rxn_rows):
        row_modules = list(map(str.strip, row[column_rxn_module].value.split(','))) # Split modules into list and trim whitespace
        
        if row[excel_book._column_reaction_full_name].value.strip() == '': # Skip empty rows
            continue
        
        if row[column_rxn_quality].value == '': # If quality missing, assume quality = 0
            row_quality = 0
        else:
            try: # Try and parse quality to an int
                row_quality = int(row[column_rxn_quality].value)
            except ValueError:
                raise TypeError('Reaction {}\'s quality must be an int.'.format(row[excel_book._column_reaction_full_name].value))
            
        if row_quality < min_quality: # If less than minimum quality skip row
            logger.debug('Skipping reaction row {} (reaction quality: {}, min quality: {})'.format(num, row_quality, min_quality))
            continue

        if not modules == [] and not any(item in row_modules for item in modules): # If row module isnt in modules list, skip row
            logger.debug('Skipping reaction row {} (module not present)'.format(num))
            continue
        else:
            rxn_filtered.append({'row_num': num, 'name': row[excel_book._column_reaction_full_name].value})
        
    # Filter rxn rows by module tag and min quality
    con_rows = [row for row in con_sheet.get_rows()][2:]
    con_filtered = []

    for num, row in enumerate(con_rows):
        row_modules = list(map(str.strip, row[column_con_module].value.split(','))) # Split modules into list and trim whitespace

        if row[excel_book._column_contingency_target].value.strip() == '': # Skip empty rows
            continue

        if row[column_con_quality].value == '': # If quality missing, assume quality = 0
            row_quality = 0
        else:
            try: # Try and parse quality to an int
                row_quality = int(row[column_con_quality].value)
            except ValueError:
                raise TypeError('Contingency {}\'s quality must be an int.'.format(num))
            
        if row_quality < min_quality: # If less than minimum quality skip row
            logger.debug('Skipping contingency row {} (reaction quality: {}, min quality: {})'.format(num, row_quality, min_quality))
            continue

        if not modules == [] and not any(item in row_modules for item in modules): # If row module isnt in modules list, skip row
            logger.debug('Skipping contingency row {} (module not present)'.format(num))
            continue
        else:
            con_filtered.append({'row_num': num, 'target': row[excel_book._column_contingency_target].value.strip(),
                                 'modifier': row[excel_book._column_contingency_modifier].value.strip(),
                                 'type': row[excel_book._column_contingency_type].value.strip()})

    ### STEP 1.5: Make sure boolean contingencies have inputs; if not, raise an error
    boolean_modifiers = [con['modifier'] for con in con_filtered if re.match(BOOLEAN_CONTINGENCY_REGEX, con['modifier'])]
    boolean_targets = [con['modifier'] for con in con_filtered if re.match(BOOLEAN_CONTINGENCY_REGEX, con['modifier'])]

    for modifier in boolean_modifiers:
        if not modifier in boolean_targets:
            raise ValueError('Boolean gate {} has no inputs in filtered contingencies'.format(modifier))

    ### STEP 2 (New method): Map modifiers to rxncon states
    required_components = []
    for con in con_filtered:
        con['modifier'] = re.sub('#.*', '', con['modifier'])

        if re.match(BOOLEAN_CONTINGENCY_REGEX, con['modifier']): # If contingency modifier is a boolean state, skip it
            con['modifier_states'] = []
            continue

        orig_modifier_state = state_from_str(con['modifier']) # Get the state based on the string
        modifier_states = [state for state in rxncon_system.states if state.is_subset_of(orig_modifier_state)] # Find the states in the system that state refers to
        
        if not modifier_states:
            raise ValueError('Could not match contingency {} to a state in the rxncon system'.format(con['modifier']))

        con['modifier_states'] = modifier_states
        for state in modifier_states:
            required_components.extend(state.components)

    ### STEP 3 (New method): For each filtered contingency (ie a state), add reaction(s) that produce(s) state to list (also add target reactions)
    def unsplit_bidirectional_reaction_str(rxn_str: str) -> str:
        for verb in BIDIRECTIONAL_REACTIONS:
            if ('_{}+_'.format(verb).lower() in rxn_str.lower() or
                    '_{}-_'.format(verb).lower() in rxn_str.lower()): # If reaction is in format _verb+_ or _verb-_
                new_rxn_str = re.sub('(?i)_{}._'.format(verb), '_{}_'.format(verb), rxn_str) # Replace with _verb_
                return new_rxn_str

        return rxn_str

    required_components = list(set(required_components)) # Make sure there are no duplicate components
    required_reactions = []
    for con in con_filtered:
        con_reactions = [] # List to store reactions producing/synthesizing contingency reactions
        con_required_reactions = [] # List to store which reactions will be added to satisfy contingency

        # If target not a global state or boolean node (i.e. target is a reaction), add it to list of required reactions
        if not (re.match(BOOLEAN_CONTINGENCY_REGEX, con['target']) or re.match(GLOBAL_STATE_REGEX, con['target'])):
            required_reactions.append(unsplit_bidirectional_reaction_str(con['target']))

        for state in con['modifier_states']: # For each state contingency modifier matched to
            for reaction in rxncon_system.reactions:
                if (any([produced_state.is_subset_of(state) for produced_state in reaction.produced_states]) or # If a reaction produces the state
                        any([synthesised_state.is_subset_of(state) for synthesised_state in reaction.synthesised_states])): # or synthesizes the state
                    con_reactions.append(reaction) # Add it to the list of reactions for the contingency
        
        # Try and find if reaction producing state already in filtered reactions
        for reaction in con_reactions: 
            rxn_name = unsplit_bidirectional_reaction_str(reaction.name) # Turn bidirectional reactions like ppi+ or ppi- into ppi
            rxn_name_no_domains = re.sub(r'_\[.*?\]', '', rxn_name) # Also check name of reaction without domains
            if [filtered_reaction for filtered_reaction in rxn_filtered if (
                                                                            filtered_reaction['name'] == rxn_name or
                                                                            filtered_reaction['name'] == rxn_name_no_domains
                                                                           )]:
                con_required_reactions.append(rxn_name)
        
        # If an already filtered reaction wasn't found, first try adding reaction involving only required components
        if not con_required_reactions: 
            for reaction in con_reactions:
                rxn_name = unsplit_bidirectional_reaction_str(reaction.name) # Turn bidirectional reactions like ppi+ or ppi- into ppi
                if all(component in required_components for component in reaction.components): # Check that all reaction components are in required_components
                    logger.warning('Adding reaction {} to satisfy contingency {} {} {}'.format(rxn_name, con['target'], con['type'], con['modifier'])) 
                    con_required_reactions.append(rxn_name)

        # If a reaction still wasn't found, add all reactions producing the contingency modifier
        if not con_required_reactions:
            for reaction in con_reactions:
                rxn_name = unsplit_bidirectional_reaction_str(reaction.name)
                logger.warning('Adding reaction {} to satisfy contingency {} {} {}'.format(rxn_name, con['target'], con['type'], con['modifier'])) 
                con_required_reactions.append(rxn_name)

        required_reactions.extend(con_required_reactions)

    ### STEP 4: Check if required reactions already in rxn_filtered and add them if missing
    all_rxn_names = rxn_sheet.col_values(excel_book._column_reaction_full_name)[2:] # Get list of all reaction names
    rxn_filtered_names = [d['name'] for d in rxn_filtered]
    required_reactions = list(set(required_reactions)) # Make sure there are no duplicate reactions
    added_reactions = []
    for required_rxn_name in required_reactions:
        # Check if reaction exists in rxn_filtered and skip it if it does
        required_rxn_name_no_domains = re.sub(r'_\[.*?\]', '', required_rxn_name) # Also check name of reaction without domains
        if required_rxn_name in rxn_filtered_names or required_rxn_name_no_domains in rxn_filtered_names:
            continue

        # If node not in rxn_filtered, try and find the matching reaction and add it to rxn_filtered
        for num, rxn_name in enumerate(all_rxn_names):
            if required_rxn_name == rxn_name or required_rxn_name_no_domains == rxn_name:
                logger.warning('Added required reaction {} to filtered reactions.'.format(rxn_name))
                added_reactions.append({'name': rxn_name, 'row_num': num})

        # If still not found, try stripping domains from both sides
        for num, rxn_name in enumerate(all_rxn_names):
            if required_rxn_name_no_domains == re.sub(r'_\[.*?\]', '', rxn_name):
                logger.warning('Added required reaction {} to filtered reactions.'.format(rxn_name))
                added_reactions.append({'name': rxn_name, 'row_num': num})

        if not added_reactions:
            raise ValueError('Unable to find reaction {}'.format(required_rxn_name))

    rxn_filtered.extend(added_reactions)

    ### STEP 5: Write updated lists to excel file. This has to be done with openpyxl as xlrd/xlwt don't really support xlsx
    # NOTE: openpyxl is 1-indexed, not 0-indexed like xlrd
    wb = load_workbook(filename = excel_filename, data_only = True)
    rxn_open_sheet = wb['ReactionList']
    con_open_sheet = wb['ContingencyList']

    # Delete the old unfiltered rows. This is more efficient than the delete_rows method
    for row in rxn_open_sheet.iter_rows(min_row=3, max_row=(3+len(rxn_rows))):
        for col in range(0,12):
            row[col].value = None

    for row in con_open_sheet.iter_rows(min_row=3, max_row=(3+len(con_rows))):
        for col in range(0,8):
            row[col].value = None 
    

    # Add the filtered rows to workbook
    rxn_filtered_row_nums = [d['row_num'] for d in rxn_filtered]
    for row_num, filtered_row_num in enumerate(rxn_filtered_row_nums, 3):
        for col_num, val in enumerate(rxn_sheet.row_values(filtered_row_num+2), 1):
            rxn_open_sheet.cell(row_num, col_num).value = val

    con_filtered_row_nums = [d['row_num'] for d in con_filtered]
    for row_num, filtered_row_num in enumerate(con_filtered_row_nums, 3):
        for col_num, val in enumerate(con_sheet.row_values(filtered_row_num+2), 1):
            con_open_sheet.cell(row_num, col_num).value = val

    # Write to file
    wb.save(output)
    print('Wrote filtered rxncon file to {}'.format(output))



@click.command()
@click.option('--file', required=True,
              help='Name of input file.')
@click.option('--output', default=None,
              help='Name of output file. Default: modules.xlsx')
@click.option('--modules', default='',
              help='Comma-separated list of modules to be extracted from input file.')
@click.option('--quality', default=0, type=click.INT,
              help='Minimum quality of rules to be kept. Default: 0')
@click_log.simple_verbosity_option(default='WARNING')
@click_log.init()
def run(file, output, modules, quality):
    modules = list(map(str.strip, modules.split(','))) # Split modules into list and trim whitespace
    modules = list(filter(None, modules)) # Filter out blank modules
    extract_modules(file, output, modules, quality)


def setup_logging_colors():
    click_log.ColorFormatter.colors = {
        'error': dict(fg='red'),
        'exception': dict(fg='red'),
        'critical': dict(fg='red'),
        'debug': dict(fg='yellow'),
        'warning': dict(fg='yellow'),
        'info': dict(fg='yellow')
    }

    def format(self, record):
        if not record.exc_info:
            level = record.levelname.lower()
            if level in self.colors:
                padding_size = 7  # Assume just INFO / DEBUG entries.

                prefix = click.style('{}: '.format(level).ljust(padding_size),
                                     **self.colors[level])

                prefix += click.style('{} '.format(record.name), fg='blue')

                msg = record.msg
                if isinstance(msg, bytes):
                    msg = msg.decode(sys.getfilesystemencoding(),
                                     'replace')
                elif not isinstance(msg, str):
                    msg = str(msg)
                record.msg = '\n'.join(prefix + x for x in msg.splitlines())

        return logging.Formatter.format(self, record)

    click_log.ColorFormatter.format = format


if __name__ == '__main__':
    try:
        setup_logging_colors()
        run()
    except Exception as e:
        print('ERROR: {}\n{}\nPlease re-run this command with the \'-v DEBUG\' option.'.format(type(e), e))

